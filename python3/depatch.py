from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime

import numpy as np
import sys

from schema import DimProject, DimEmployee
from util import MonthConverter
from util import isNullVal

class Depatcher:
  def __init__(self, source, target):
    self.current_time = MonthConverter.month_to_int(int(datetime.now().strftime('%m')))
    self.seperator = '|'
    self.wb = load_workbook(source)
    self.target = target
    self.dim_project = self.wb['DimProject'] if 'DimProject' in self.wb else self.wb.create_sheet('DimProject')
    self.dim_employee = self.wb['DimEmployee'] if 'DimEmployee' in self.wb else self.wb.create_sheet('DimEmployee')
    self.est = self.wb['est'] if 'est' in self.wb else self.wb.create_sheet('est')
    self.act = self.wb['act'] if 'act' in self.wb else self.wb.create_sheet('act')
  def exec(self):
    dim_project_intern = self.readDimProject(self.dim_project)
    dim_employee_intern = self.readDimEmployee(self.dim_employee)
    project_map = dict([(f'{dim_project_intern[i].ma_or_project}{self.seperator}{dim_project_intern[i].category}{self.seperator}{dim_project_intern[i].name}', i) for i in range(len(dim_project_intern))])
    employee_map = dict([(f'{dim_employee_intern[i].category}{self.seperator}{dim_employee_intern[i].role}{self.seperator}{dim_employee_intern[i].itcode}', i) for i in range(len(dim_employee_intern))])
    est_intern = self.readCross('est', project_map, employee_map)
    act_intern = self.readCross('act', project_map, employee_map)
    for i in range(4, 16):
      for j in range(np.size(act_intern[i], 0)):
        if len([e for e in act_intern[i][j] if e is not None]) == 0:
          continue
        else:
          project_act_expend = sum(np.array([e if e is not None else 0.0 for e in act_intern[i][j]]) * np.array([e.rate[i] if e.rate[i] is not None else 0.0 for e in dim_employee_intern]))
          dim_project_intern[j].updateBudgetByTime(i, project_act_expend)
    self.updateDimProject(dim_project_intern)
    self.updateDimEmployee(dim_employee_intern)
    for i in range(self.current_time, 16):
      self.depatch(i, est_intern, dim_project_intern, dim_employee_intern)
    self.updateCross('est', est_intern, dim_project_intern, dim_employee_intern)
    self.updateCross('act', act_intern, dim_project_intern, dim_employee_intern)
    self.wb.save(self.target)

  def depatch(self, time, est_intern, dim_project_intern, dim_employee_intern):
    current_est = est_intern[time]
    if len([current_est[i, j] for i in range(np.size(current_est, 0)) for j in range(np.size(current_est, 1)) if current_est[i, j] is not None]) == 0:
      print(f'initialized est in month {MonthConverter.int_to_month(time)}')
      return
    non_ma_index = [
      i
      for i in range(np.size(current_est, 0))
      if dim_project_intern[i].ma_or_project != 'MA'
    ]
    for r in non_ma_index:
      if dim_project_intern[r].budget_by_time[time] is None:
        current_est[r] = [None] * np.size(current_est, 1)
        continue
      employee_index = [
        i
        for i in range(np.size(current_est, 1))
        if current_est[r, i] is not None and dim_project_intern[r].category == dim_employee_intern[i].category
      ]
      if len(employee_index) == 0:
        raise Exception(f'no role depatch for project {dim_project_intern[r].name} in month {MonthConverter.int_to_month(time)}')
      
      employee_rate = np.array([dim_employee_intern[i].rate[time] for i in employee_index])
      if None in employee_rate:
        raise Exception(f'unknown employee rate for project {dim_project_intern[r].name} in month {MonthConverter.int_to_month(time)}')
      budget_depatch_sum = sum(current_est[r, employee_index] * employee_rate)
      if budget_depatch_sum >= 0.95 * dim_project_intern[r].budget_by_time[time] and budget_depatch_sum <= dim_project_intern[r].budget_by_time[time]:
        continue

      current_est[r] = [None] * np.size(current_est, 1)

      employee_md_total = np.array([dim_employee_intern[i].md[time] for i in employee_index])
      if None in employee_md_total:
        raise Exception(f'unknown employee md for project {dim_project_intern[r].name} in month {MonthConverter.int_to_month(time)}')

      employee_md_depatched = np.array([sum([current_est[j, i] for j in non_ma_index if current_est[j, i] is not None]) for i in employee_index])
      employee_md_remain = np.clip(employee_md_total - employee_md_depatched, 0, sys.maxsize)
      employee_roles = [dim_employee_intern[i].role for i in employee_index]

      employee_depatch = np.array([0.0] * len(employee_index))
      project_remain = dim_project_intern[r].budget_by_time[time]
      employee_md_remain_sort_index = list(np.flip(np.argsort(employee_md_remain)))
      employee_md_remain_sort_index_with_role = []
      employee_roles_buff = set()
      while len(employee_md_remain_sort_index) > 0:
        if len(employee_roles_buff) == 0:
          employee_roles_buff = set(employee_roles)
        employee_md_remain_sort_index_len = len(employee_md_remain_sort_index)
        for i in range(employee_md_remain_sort_index_len):
          if employee_roles[employee_md_remain_sort_index[i]] in employee_roles_buff:
            employee_roles_buff.remove(employee_roles[employee_md_remain_sort_index[i]])
            employee_md_remain_sort_index_with_role.append(employee_md_remain_sort_index.pop(i))
            break
        if i == (employee_md_remain_sort_index_len - 1):
          employee_roles_buff = set(employee_roles)

      for i in employee_md_remain_sort_index_with_role:
        if project_remain >= employee_md_remain[i] * employee_rate[i]:
          employee_depatch[i] = employee_md_remain[i]
          project_remain -= employee_md_remain[i] * employee_rate[i]
        elif project_remain > 0:
          employee_depatch[i] = project_remain / employee_rate[i]
          project_remain = 0
        else:
          break
      current_est[r, employee_index] = employee_depatch

      if project_remain > 0:
        employee_index_expand = [
          i
          for i in range(np.size(current_est, 1))
          if dim_employee_intern[i].role in employee_roles and
             dim_project_intern[r].category == dim_employee_intern[i].category and
             dim_employee_intern[i].rate[time] is not None and
             dim_employee_intern[i].md[time] is not None and
             i not in employee_index
        ]
        employee_rate_expand = np.array([dim_employee_intern[i].rate[time] for i in employee_index_expand])
        employee_md_total_expand = np.array([dim_employee_intern[i].md[time] for i in employee_index_expand])
        employee_md_depatched_expand = np.array([sum([current_est[j, i] for j in non_ma_index if current_est[j, i] is not None]) for i in employee_index_expand])
        employee_md_remain_expand = np.clip(employee_md_total_expand - employee_md_depatched_expand, 0, sys.maxsize)
        employee_roles_expand = [dim_employee_intern[i].role for i in employee_index_expand]

        employee_depatch_expand = np.array([None] * len(employee_index_expand))
        employee_md_remain_sort_index_expand = list(np.flip(np.argsort(employee_md_remain_expand)))
        employee_md_remain_sort_index_with_role_expand = []
        employee_roles_expand_buff = set()

        while len(employee_md_remain_sort_index_expand) > 0:
          if len(employee_roles_expand_buff) == 0:
            employee_roles_expand_buff = set(employee_roles_expand)
          employee_md_remain_sort_index_expand_len = len(employee_md_remain_sort_index_expand)
          for i in range(employee_md_remain_sort_index_expand_len):
            if employee_roles_expand[employee_md_remain_sort_index_expand[i]] in employee_roles_expand_buff:
              employee_roles_expand_buff.remove(employee_roles_expand[employee_md_remain_sort_index_expand[i]])
              employee_md_remain_sort_index_with_role_expand.append(employee_md_remain_sort_index_expand.pop(i))
              break
          if i == (employee_md_remain_sort_index_expand_len - 1):
            employee_roles_expand_buff = set(employee_roles_expand)
        for i in employee_md_remain_sort_index_with_role_expand:
          if project_remain >= employee_md_remain_expand[i] * employee_rate_expand[i]:
            employee_depatch_expand[i] = employee_md_remain_expand[i]
            project_remain -= employee_md_remain_expand[i] * employee_rate_expand[i]
          elif project_remain > 0:
            employee_depatch_expand[i] = project_remain / employee_rate_expand[i]
            project_remain = 0
          else:
            break
        current_est[r, employee_index_expand] = employee_depatch_expand

        if project_remain > 0:
          raise Exception(f'cannot expend budget of project {dim_project_intern[r].name} in month {MonthConverter.int_to_month(time)} after expand on roles')

    ma_index = [
      i
      for i in range(np.size(current_est, 0))
      if dim_project_intern[i].ma_or_project == 'MA'
    ]
    ma_project_remain = [dim_project_intern[i].budget_by_time[time] for i in ma_index]
    for c in range(np.size(current_est, 1)):
      md_total = dim_employee_intern[c].md[time]
      rate_c = dim_employee_intern[c].rate[time]
      if rate_c is None:
        current_est[ma_index, c] = [None] * len(ma_index)
        continue
      md_depatched = sum([current_est[i, c] for i in non_ma_index if current_est[i, c] is not None])
      md_remain = max(0.0, md_total - md_depatched)
      if len([current_est[i, c] for i in ma_index if dim_project_intern[i].category == dim_employee_intern[c].category and current_est[i, c] is not None]) == 0:
        ma_index_selected = np.argmax([ma_project_remain[i] if dim_project_intern[i].category == dim_employee_intern[c].category else -sys.maxsize for i in ma_index])
        current_est[ma_index, c] = [None] * len(ma_index)
        current_est[ma_index[ma_index_selected], c] = md_remain if md_remain > 0 else None
      else:
        ma_index_selected = np.argmax([current_est[i, c] if dim_project_intern[i].category == dim_employee_intern[c].category and current_est[i, c] is not None else -sys.maxsize for i in ma_index])
        current_est[ma_index, c] = [None] * len(ma_index)
        current_est[ma_index[ma_index_selected], c] = md_remain
      ma_project_remain[ma_index_selected] = ma_project_remain[ma_index_selected] - md_remain * rate_c


  def updateCross(self, est_or_act, cross_intern, dim_project_intern, dim_employee_intern):
    if est_or_act == 'est':
      sheet = self.est
    elif est_or_act == 'act':
      sheet = self.act
    else:
      raise Exception(f'unknown cross type {est_or_act}')
    for i in range(len(dim_project_intern)):
      sheet.cell(row = 1, column = i + 3).value = f'{dim_project_intern[i].ma_or_project}{self.seperator}{dim_project_intern[i].category}{self.seperator}{dim_project_intern[i].name}'
    for i in range(len(dim_employee_intern)):
      for j in range(4, 16):
        sheet.cell(row = i * 12 + j - 2, column = 1).value = MonthConverter.int_to_month(j)
        sheet.cell(row = i * 12 + j - 2, column = 2).value = f'{dim_employee_intern[i].category}{self.seperator}{dim_employee_intern[i].role}{self.seperator}{dim_employee_intern[i].itcode}'
    for i in range(len(dim_project_intern)):
      for j in range(len(dim_employee_intern)):
        for k in range(4, 16):
          sheet.cell(row = j * 12 + k - 2, column = i + 3).value = cross_intern[k][i, j]

  def readCross(self, est_or_act, project_map, employee_map):
    cross_intern = dict([(i, np.array([[None] * len(employee_map)] * len(project_map))) for i in range(4, 16)])
    if est_or_act == 'est':
      sheet = self.est
    elif est_or_act == 'act':
      sheet = self.act
    else:
      raise Exception(f'unknown cross type {est_or_act}')
    for c in range(3, sheet.max_column + 1):
      for r in range(2, sheet.max_row + 1):
        project = sheet.cell(row = 1, column = c).value
        employee = sheet.cell(row = r, column = 2).value
        time = MonthConverter.month_to_int(int(sheet.cell(row = r, column = 1).value))
        value = sheet.cell(row = r, column = c).value
        if not isNullVal(value):
          cross_intern[time][project_map[project], employee_map[employee]] = float(value)
    return cross_intern


  def readDimProject(self, dim_project):
    title_column_map = {}
    dim_project_intern = []
    for i in range(1, dim_project.max_column + 1):
      title = dim_project.cell(row = 1, column = i).value
      if title is None:
        break
      else:
        title_column_map[title] = i
    for r in range(2, dim_project.max_row + 1):
      row = DimProject(
        name = dim_project.cell(row = r, column = title_column_map['name']).value,
        status = dim_project.cell(row = r, column = title_column_map['status']).value,
        ma_or_project = dim_project.cell(row = r, column = title_column_map['ma_or_project']).value,
        category = dim_project.cell(row = r, column = title_column_map['category']).value,
        budget = dim_project.cell(row = r, column = title_column_map['budget']).value,
        month_start = dim_project.cell(row = r, column = title_column_map['month_start']).value,
        month_end = dim_project.cell(row = r, column = title_column_map['month_end']).value,
        budget_by_month = {k: dim_project.cell(row = r, column = v).value for k, v in title_column_map.items() if k not in ['name', 'status', 'ma_or_project', 'category', 'budget', 'month_start', 'month_end', 'remain']}
      )
      dim_project_intern.append(row)
      dim_project_intern.sort(key = lambda e: f'{e.ma_or_project}{e.category}{e.name}')
    return dim_project_intern
  def updateDimProject(self, dim_project_intern):
    self.dim_project.cell(row = 1, column = 1).value = 'name'
    self.dim_project.cell(row = 1, column = 2).value = 'status'
    self.dim_project.cell(row = 1, column = 3).value = 'ma_or_project'
    self.dim_project.cell(row = 1, column = 4).value = 'category'
    self.dim_project.cell(row = 1, column = 5).value = 'budget'
    self.dim_project.cell(row = 1, column = 6).value = 'month_start'
    self.dim_project.cell(row = 1, column = 7).value = 'month_end'
    self.dim_project.cell(row = 1, column = 8).value = 'remain'
    for i in range(4, 16):
      self.dim_project.cell(row = 1, column = 5 + i).value = f'budget_by_month_{MonthConverter.int_to_month(i)}'
    for r in range(len(dim_project_intern)):
      self.dim_project.cell(row = r + 2, column = 1).value = dim_project_intern[r].name
      self.dim_project.cell(row = r + 2, column = 2).value = dim_project_intern[r].status
      self.dim_project.cell(row = r + 2, column = 3).value = dim_project_intern[r].ma_or_project
      self.dim_project.cell(row = r + 2, column = 4).value = dim_project_intern[r].category
      self.dim_project.cell(row = r + 2, column = 5).value = dim_project_intern[r].budget
      self.dim_project.cell(row = r + 2, column = 6).value = MonthConverter.int_to_month(dim_project_intern[r].time_start)
      self.dim_project.cell(row = r + 2, column = 7).value = MonthConverter.int_to_month(dim_project_intern[r].time_end)
      self.dim_project.cell(row = r + 2, column = 8).value = dim_project_intern[r].remain
      for c in range(4, 16):
        self.dim_project.cell(row = r + 2, column = c + 5).value = dim_project_intern[r].budget_by_time[c]

  def readDimEmployee(self, dim_employee):
    title_column_map = {}
    dim_employee_intern = []
    for i in range(1, dim_employee.max_column + 1):
      title = dim_employee.cell(row = 1, column = i).value
      if title is None:
        break
      else:
        title_column_map[title] = i
    for r in range(2, dim_employee.max_row + 1):
      row = DimEmployee(
        itcode = dim_employee.cell(row = r, column = title_column_map['itcode']).value,
        category = dim_employee.cell(row = r, column = title_column_map['category']).value,
        role = dim_employee.cell(row = r, column = title_column_map['role']).value,
        rate_and_md = {k: dim_employee.cell(row = r, column = v).value  for k, v in title_column_map.items() if k not in ['itcode', 'category', 'role']}
      )
      dim_employee_intern.append(row)
      dim_employee_intern.sort(key = lambda e: f'{e.category}{e.role}{e.itcode}')
    return dim_employee_intern

  def updateDimEmployee(self, dim_employee_intern):
    self.dim_employee.cell(row = 1, column = 1).value = 'itcode'
    self.dim_employee.cell(row = 1, column = 2).value = 'category'
    self.dim_employee.cell(row = 1, column = 3).value = 'role'
    for i in range(4, 16):
      self.dim_employee.cell(row = 1, column = i).value = f'rate_{MonthConverter.int_to_month(i)}'
      self.dim_employee.cell(row = 1, column = 12 + i).value = f'md_{MonthConverter.int_to_month(i)}'
    for r in range(len(dim_employee_intern)):
      self.dim_employee.cell(row = r + 2, column = 1).value = dim_employee_intern[r].itcode
      self.dim_employee.cell(row = r + 2, column = 2).value = dim_employee_intern[r].category
      self.dim_employee.cell(row = r + 2, column = 3).value = dim_employee_intern[r].role
      for c in range(4, 16):
        self.dim_employee.cell(row = r + 2, column = c).value = dim_employee_intern[r].rate[c]
        self.dim_employee.cell(row = r + 2, column = 12 + c).value = dim_employee_intern[r].md[c]
